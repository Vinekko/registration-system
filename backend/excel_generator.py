from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from typing import List, Dict

class ExcelGenerator:
    """Genera archivos Excel desde los registros"""
    
    @staticmethod
    def generate_registros_excel(registros: List[Dict]) -> BytesIO:
        """Genera un archivo Excel con todos los registros"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros Adakademy"
        
        # Encabezados
        headers = ["ID", "Nombre", "Teléfono", "Correo", "Servicio"]
        
        # Estilos para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escribir datos
        for row, registro in enumerate(registros, 2):
            ws.cell(row=row, column=1, value=registro.get("id", ""))
            ws.cell(row=row, column=2, value=registro.get("nombre", ""))
            ws.cell(row=row, column=3, value=registro.get("telefono", ""))
            ws.cell(row=row, column=4, value=registro.get("email", ""))
            ws.cell(row=row, column=5, value=registro.get("servicio", ""))
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 50
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output